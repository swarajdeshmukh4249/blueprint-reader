"""
Export BOQ to CSV, Excel, and PDF (letterhead-ready).
"""

from __future__ import annotations
import csv
import io
from datetime import datetime
from typing import Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False


def _rows(items: list[dict]) -> list[list[Any]]:
    header = ["#", "Description", "Unit", "Qty", "Rate (₹)", "Amount (₹)", "Category"]
    rows = [header]
    for i, it in enumerate(items, 1):
        rows.append([
            i,
            it.get("description") or it.get("item") or "",
            it.get("unit") or "",
            it.get("quantity") or it.get("qty") or 0,
            it.get("rate") or 0,
            it.get("amount") or 0,
            it.get("category") or "",
        ])
    return rows


def export_csv(items: list[dict], summary: Optional[dict] = None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in _rows(items):
        w.writerow(row)
    if summary:
        w.writerow([])
        w.writerow(["Subtotal (material)", summary.get("material_subtotal", "")])
        w.writerow(["Subtotal (labour)", summary.get("labour_subtotal", "")])
        w.writerow(["GST", summary.get("gst_amount", "")])
        w.writerow(["Grand Total", summary.get("grand_total", "")])
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(
    items: list[dict],
    summary: Optional[dict] = None,
    company_name: str = "Blueprint Reader BOQ",
) -> bytes:
    if not OPENPYXL:
        raise RuntimeError("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws["A1"] = company_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    start = 4
    for r_idx, row in enumerate(_rows(items), start):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start:
                cell.font = Font(bold=True)
    if summary:
        r = start + len(items) + 2
        for label, key in [
            ("Material", "material_subtotal"),
            ("Labour", "labour_subtotal"),
            ("GST", "gst_amount"),
            ("Grand Total", "grand_total"),
        ]:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=summary.get(key, 0))
            r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(
    items: list[dict],
    summary: Optional[dict] = None,
    company_name: str = "Blueprint Reader BOQ",
    letterhead_line: str = "",
) -> bytes:
    if not FPDF_AVAILABLE:
        raise RuntimeError("fpdf2 not installed")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, company_name, ln=True)
    if letterhead_line:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, letterhead_line, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, datetime.utcnow().strftime("Date: %Y-%m-%d"), ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 8)
    cols = ["#", "Description", "Unit", "Qty", "Rate", "Amount"]
    widths = [8, 70, 18, 18, 28, 28]
    for w, c in zip(widths, cols):
        pdf.cell(w, 7, c, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for i, it in enumerate(items, 1):
        row = [
            str(i),
            (it.get("description") or it.get("item") or "")[:45],
            str(it.get("unit") or "")[:8],
            str(it.get("quantity") or it.get("qty") or 0),
            str(it.get("rate") or 0),
            str(it.get("amount") or 0),
        ]
        for w, val in zip(widths, row):
            pdf.cell(w, 6, val, border=1)
        pdf.ln()

    if summary:
        pdf.ln(4)
        pdf.set_font("Helvetica", "", 9)
        for label, key in [
            ("Material subtotal", "material_subtotal"),
            ("Labour subtotal", "labour_subtotal"),
            ("GST", "gst_amount"),
            ("Grand Total", "grand_total"),
        ]:
            pdf.cell(0, 6, f"{label}: ₹ {summary.get(key, 0):,.2f}", ln=True)

    out = pdf.output()
    return out if isinstance(out, bytes) else out.encode("latin-1")
